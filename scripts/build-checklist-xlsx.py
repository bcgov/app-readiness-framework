# -*- coding: utf-8 -*-
"""Generate the professional Excel checklist from the single source of truth.

Source : docs/checklist/checklist-items.csv   (edit this in Excel)
Output : Application-Readiness-Checklist.xlsx  (repo root, or pass a path)

Usage:
    pip install openpyxl
    python scripts/build-checklist-xlsx.py                 # writes to repo root
    python scripts/build-checklist-xlsx.py "C:/path/out.xlsx"

The CSV drives the live tool AND this spreadsheet, so they stay in sync. Edit the
CSV, open a PR, and re-run this script to refresh the Excel copy.
"""
import csv, os, sys
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
CSV_PATH = os.path.join(REPO, "docs", "checklist", "checklist-items.csv")
OUT = sys.argv[1] if len(sys.argv) > 1 else os.path.join(REPO, "Application-Readiness-Checklist.xlsx")

NAVY = "003366"; NAVY2 = "16324F"; GOLD = "FCBA19"; INK = "1B2733"

# A static spreadsheet can't be platform-aware; show the OpenShift defaults for the tokens.
TOKENS = {"{{MONITOR}}": "Sysdig", "{{LOGS}}": "the Hive", "{{SECRETS}}": "Vault"}
def resolve(s):
    for k, v in TOKENS.items():
        s = s.replace(k, v)
    return s

# The CSV "More info" paths are relative to the checklist page; resolve to live URLs.
LINK_BASE = "https://bcgov.github.io/app-readiness-framework/checklist/checklist-generator/"
def linkify(raw):
    raw = (raw or "").strip()
    if not raw:
        return ""
    return raw if raw.startswith("http") else urljoin(LINK_BASE, raw)

# section -> tint (gate comes from the CSV); fallback tint for any new section.
TINT = {
    "Design & decisions": "E6F1FB", "Build & pipeline": "EAF3DE",
    "Resilience & performance": "EEEDFE", "Data & DR": "E1F5EE",
    "Observability": "FAEEDA", "Security & access": "FBEAF0",
    "Operability & support": "F1EFE8", "Vendor (contractual)": "FAECE7",
}
def tint(sec):
    return TINT.get(sec, "EFEFEF")

# ---- read the CSV (the source of truth) ----
ROWS = []
with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
    for row in csv.DictReader(f):
        if not (row.get("Item") or "").strip():
            continue
        ROWS.append({
            "sec": row["Section"].strip(), "gate": row["Gate"].strip(),
            "item": resolve(row["Item"].strip()), "why": resolve(row["Why"].strip()),
            "ev": resolve(row["Evidence"].strip()), "covers": resolve((row.get("Covers") or "").strip()),
            "t1": row["Tier 1"].strip(), "t2": row["Tier 2"].strip(), "t3": row["Tier 3"].strip(),
            "ap": (row.get("Applies to") or "All").strip(), "link": linkify(row.get("More info")),
        })

wb = Workbook()
thin = Side(style="thin", color="D6D6D6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------- Cover sheet ----------
cov = wb.active; cov.title = "Read me"; cov.sheet_view.showGridLines = False
cov.merge_cells("A1:B1")
c = cov["A1"]; c.value = "Application Readiness Checklist"
c.font = Font(bold=True, size=20, color="FFFFFF")
c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
cov["A1"].fill = PatternFill("solid", fgColor=NAVY); cov.row_dimensions[1].height = 42
cov.merge_cells("A2:B2"); cov["A2"].fill = PatternFill("solid", fgColor=GOLD); cov.row_dimensions[2].height = 5
def line(row, txt, bold=False, size=10, color=INK, indent=1):
    cov.merge_cells(f"A{row}:B{row}")
    cell = cov.cell(row, 1, txt)
    cell.font = Font(bold=bold, size=size, color=color)
    cell.alignment = Alignment(wrap_text=True, vertical="top", indent=indent)
r = 4
for t, b, s, col in [
    ("Guardrails for building, hardening, and handing over supportable, secure, resilient applications.", False, 11, "555555"),
    ("For internal teams, contractors and vendors. Right-sized by criticality tier.", False, 11, "555555"),
    ("", False, 10, INK),
    ("The four gates", True, 13, NAVY),
    ("G1 Design  \u2192  G2 Build  \u2192  G3 Readiness  \u2192  G4 Operate.  Each item below is checked at its gate.", False, 10, INK),
    ("", False, 10, INK),
    ("Criticality tiers  (how critical the application is \u2014 drives the obligation)", True, 13, NAVY),
    ("Tier 1 \u2014 Mission-critical    \u00b7    Tier 2 \u2014 Business-important    \u00b7    Tier 3 \u2014 Supporting", False, 10, INK),
    ("", False, 10, INK),
    ("Obligation key", True, 13, NAVY),
    ("Must = required to pass the gate    \u00b7    Should = strongly recommended (justify deviation in an ADR)", False, 10, INK),
    ("Optional = situational    \u00b7    N/A = does not apply at that tier", False, 10, INK),
    ("", False, 10, INK),
    ("How to use", True, 13, NAVY),
    ("1. Pick your tier and platform.   2. Work each applicable item and attach the evidence.   3. Use the 'More info' link for detail.", False, 10, INK),
    ("", False, 10, INK),
    ("Principle", True, 13, NAVY),
    ("'More info' links point to EXISTING BC Gov / developer.gov.bc.ca pages. This checklist does not author or maintain guidance \u2014 it points to it.", False, 10, "555555"),
    ("Tool names shown (Sysdig, the Hive, Vault) are the OpenShift defaults; Salesforce / public cloud use their equivalents.", False, 10, "555555"),
    ("", False, 10, INK),
    (f"{len(ROWS)} items \u00b7 consolidated from 70 \u00b7 generated from checklist-items.csv (the single source of truth) \u00b7 see the Dictionary tab.", False, 9, "888888"),
]:
    line(r, t, b, s, col); r += 1
cov.column_dimensions["A"].width = 95; cov.column_dimensions["B"].width = 4

# ---------- Checklist sheet ----------
ws = wb.create_sheet("Checklist"); ws.sheet_view.showGridLines = False
headers = ["#", "Section", "Gate", "Item", "Why (justification)", "Evidence expected", "What it covers",
           "Tier 1", "Tier 2", "Tier 3", "Applies to", "More info (link)", "Team notes"]
ws.append(headers)
hfill = PatternFill("solid", fgColor=NAVY); hfont = Font(bold=True, color="FFFFFF", size=10)
for cix in range(1, len(headers) + 1):
    cell = ws.cell(1, cix); cell.fill = hfill; cell.font = hfont; cell.border = border
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
ws.row_dimensions[1].height = 30
ob_font = {
    "Must": Font(bold=True, color="A32D2D", size=9), "Should": Font(color="854F0B", size=9),
    "Optional": Font(color="185FA5", size=9), "N/A": Font(color="AAAAAA", size=9),
}
r = 2
for i, d in enumerate(ROWS, start=1):
    vals = [i, d["sec"], d["gate"], d["item"], d["why"], d["ev"], d["covers"],
            d["t1"], d["t2"], d["t3"], d["ap"], d["link"], ""]
    for cix, v in enumerate(vals, start=1):
        cell = ws.cell(r, cix, v); cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if cix == 1: cell.alignment = Alignment(vertical="top", horizontal="center")
        if cix == 2: cell.fill = PatternFill("solid", fgColor=tint(d["sec"])); cell.font = Font(bold=True, size=9, color=NAVY2)
        if cix == 3: cell.alignment = Alignment(vertical="top", horizontal="center"); cell.font = Font(size=9, color="555555")
        if cix == 4: cell.font = Font(bold=True, size=10, color=INK)
        if cix in (8, 9, 10): cell.font = ob_font.get(v, Font(size=9)); cell.alignment = Alignment(horizontal="center", vertical="top")
        if cix == 12 and isinstance(v, str) and v.startswith("http"):
            cell.hyperlink = v; cell.value = "Open guidance \u2197"; cell.font = Font(color="185FA5", underline="single", size=9)
        elif cix == 12:
            cell.font = Font(size=9, color="555555")
    r += 1
dv = DataValidation(type="list", formula1='"Must,Should,Optional,N/A"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"H2:J{r - 1}")
widths = [4, 19, 6, 34, 40, 42, 40, 7, 7, 7, 13, 24, 24]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = "D2"; ws.auto_filter.ref = f"A1:M{r - 1}"

# ---------- Dictionary ----------
info = wb.create_sheet("Dictionary"); info.sheet_view.showGridLines = False
D = [
    ("Column dictionary", True, 14), ("", False, 10),
    ("Section \u2014 the lifecycle area the item belongs to (colour-coded).", False, 10),
    ("Gate \u2014 G1 Design \u00b7 G2 Build \u00b7 G3 Readiness (when the item is checked).", False, 10),
    ("Item \u2014 the single thing to do.", False, 10),
    ("Why (justification) \u2014 why it matters, in plain language.", False, 10),
    ("Evidence expected \u2014 what proves it's done (a link / report you attach).", False, 10),
    ("What it covers \u2014 the sub-practices this one item rolls up (kept the list short without losing detail).", False, 10),
    ("Tier 1 / Tier 2 / Tier 3 \u2014 the CRITICALITY of the app (1 = mission-critical \u2026 3 = supporting).", False, 10),
    ("     The cell value is the OBLIGATION at that tier: Must \u00b7 Should \u00b7 Optional \u00b7 N/A.", False, 10),
    ("Applies to \u2014 the PLATFORM/CONTEXT the item shows for: All \u00b7 Salesforce \u00b7 Vendor-built \u00b7 Public-facing \u00b7 OpenShift.", False, 10),
    ("More info (link) \u2014 link to an EXISTING BC Gov / dev-hub page. We do not author docs.", False, 10),
    ("Team notes \u2014 reviewer comments.", False, 10),
    ("", False, 10),
    ("Tier vs Applies-to (the common confusion)", True, 12),
    ("  Tier = how critical the APP is  (drives Must / Should / Optional).", False, 10),
    ("  Applies-to = what PLATFORM the app runs on  (drives whether the item appears at all).", False, 10),
    ("", False, 10),
    ("Source of truth", True, 12),
    ("  This workbook is generated from checklist-items.csv in the repo. Edit that CSV (in Excel), open a PR,", False, 10),
    ("  and the live tool AND this spreadsheet stay in sync. Regenerate with scripts/build-checklist-xlsx.py.", False, 10),
]
for i, (t, b, sz) in enumerate(D, start=1):
    info.cell(i, 1, t).font = Font(bold=b, size=sz, color=NAVY if b else "333333")
info.column_dimensions["A"].width = 112

try:
    wb.save(OUT); path = OUT
except PermissionError:
    base, ext = os.path.splitext(OUT)
    path = base + " (new)" + ext
    wb.save(path)
    print("NOTE: target was open/locked; wrote a copy instead.")
print("WROTE", path, "| items", len(ROWS), "| sheets", wb.sheetnames)
